from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import tempfile


def generate_excel_report(stats, sales, top_products, user_name="Пользователь"):
    """Генерация Excel отчета"""

    # Создаем workbook
    wb = Workbook()

    # Стили
    header_fill = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=12)
    title_font = Font(bold=True, size=16)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # ==================== Лист 1: Сводка ====================
    ws_summary = wb.active
    ws_summary.title = "Сводка"

    # Заголовок
    ws_summary['A1'] = "Отчет о продажах"
    ws_summary['A1'].font = title_font
    ws_summary['A2'] = f"Дата: {datetime.now().strftime('%d.%m.%Y %H:%M')}"
    ws_summary['A3'] = f"Пользователь: {user_name}"

    # Сводные показатели
    ws_summary['A5'] = "Показатель"
    ws_summary['B5'] = "Значение"
    ws_summary['A5'].fill = header_fill
    ws_summary['B5'].fill = header_fill
    ws_summary['A5'].font = header_font
    ws_summary['B5'].font = header_font

    summary_data = [
        ['Общая выручка', f"{stats['total_amount']:,.2f} ₽"],
        ['Всего продаж', stats['total_sales']],
        ['Средний чек', f"{stats['average_check']:,.2f} ₽"],
        ['Завершено', stats['completed_sales']],
        ['В ожидании', stats['pending_sales']],
        ['Отменено', stats['cancelled_sales']],
        ['Конверсия', f"{stats['conversion_rate']:.2f}%"]
    ]

    row = 6
    for label, value in summary_data:
        ws_summary[f'A{row}'] = label
        ws_summary[f'B{row}'] = value
        ws_summary[f'A{row}'].border = border
        ws_summary[f'B{row}'].border = border
        row += 1

    # Автоширина столбцов
    ws_summary.column_dimensions['A'].width = 25
    ws_summary.column_dimensions['B'].width = 20

    # ==================== Лист 2: Продажи ====================
    ws_sales = wb.create_sheet("Продажи")

    # Заголовки
    headers = ['Дата', 'Товар', 'Сумма', 'Количество', 'Статус']
    for col, header in enumerate(headers, 1):
        cell = ws_sales.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal='center')

    # Данные продаж
    status_map = {
        'completed': 'Завершено',
        'pending': 'Ожидание',
        'cancelled': 'Отменено'
    }

    for idx, sale in enumerate(sales, 2):
        date_str = datetime.fromisoformat(sale['date'].replace('Z', '+00:00')).strftime('%d.%m.%Y %H:%M')

        ws_sales[f'A{idx}'] = date_str
        ws_sales[f'B{idx}'] = sale['product_name']
        ws_sales[f'C{idx}'] = sale['amount']
        ws_sales[f'D{idx}'] = sale['quantity']
        ws_sales[f'E{idx}'] = status_map.get(sale['status'], sale['status'])

        # Границы
        for col in range(1, 6):
            ws_sales.cell(row=idx, column=col).border = border

        # Форматирование суммы
        ws_sales[f'C{idx}'].number_format = '#,##0.00 ₽'

    # Автоширина
    ws_sales.column_dimensions['A'].width = 18
    ws_sales.column_dimensions['B'].width = 40
    ws_sales.column_dimensions['C'].width = 15
    ws_sales.column_dimensions['D'].width = 12
    ws_sales.column_dimensions['E'].width = 15

    # Итоговая строка
    total_row = len(sales) + 2
    ws_sales[f'A{total_row}'] = "ИТОГО:"
    ws_sales[f'A{total_row}'].font = Font(bold=True)
    ws_sales[f'C{total_row}'] = f"=SUM(C2:C{total_row-1})"
    ws_sales[f'C{total_row}'].font = Font(bold=True)
    ws_sales[f'C{total_row}'].number_format = '#,##0.00 ₽'
    ws_sales[f'D{total_row}'] = f"=SUM(D2:D{total_row-1})"
    ws_sales[f'D{total_row}'].font = Font(bold=True)

    # ==================== Лист 3: Топ товаров ====================
    ws_top = wb.create_sheet("Топ товаров")

    # Заголовки
    top_headers = ['№', 'Товар', 'Выручка', 'Количество', 'Продаж']
    for col, header in enumerate(top_headers, 1):
        cell = ws_top.cell(row=1, column=col, value=header)
        cell.fill = PatternFill(start_color="10B981", end_color="10B981", fill_type="solid")
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal='center')

    # Данные топ товаров
    for idx, product in enumerate(top_products[:10], 2):
        ws_top[f'A{idx}'] = idx - 1
        ws_top[f'B{idx}'] = product['product_name']
        ws_top[f'C{idx}'] = product['total_amount']
        ws_top[f'D{idx}'] = product['total_quantity']
        ws_top[f'E{idx}'] = product['sales_count']

        # Границы
        for col in range(1, 6):
            ws_top.cell(row=idx, column=col).border = border

        # Форматирование
        ws_top[f'C{idx}'].number_format = '#,##0.00 ₽'

    # Автоширина
    ws_top.column_dimensions['A'].width = 5
    ws_top.column_dimensions['B'].width = 40
    ws_top.column_dimensions['C'].width = 18
    ws_top.column_dimensions['D'].width = 15
    ws_top.column_dimensions['E'].width = 12

    # ==================== Лист 4: Аналитика ====================
    ws_analytics = wb.create_sheet("Аналитика")

    # Анализ по статусам
    ws_analytics['A1'] = "Анализ по статусам"
    ws_analytics['A1'].font = title_font

    ws_analytics['A3'] = "Статус"
    ws_analytics['B3'] = "Количество"
    ws_analytics['C3'] = "Процент"
    for col in ['A3', 'B3', 'C3']:
        ws_analytics[col].fill = header_fill
        ws_analytics[col].font = header_font

    ws_analytics['A4'] = "Завершено"
    ws_analytics['B4'] = stats['completed_sales']
    ws_analytics['C4'] = f"=B4/{stats['total_sales']}*100"
    ws_analytics['C4'].number_format = '0.00"%"'

    ws_analytics['A5'] = "В ожидании"
    ws_analytics['B5'] = stats['pending_sales']
    ws_analytics['C5'] = f"=B5/{stats['total_sales']}*100"
    ws_analytics['C5'].number_format = '0.00"%"'

    ws_analytics['A6'] = "Отменено"
    ws_analytics['B6'] = stats['cancelled_sales']
    ws_analytics['C6'] = f"=B6/{stats['total_sales']}*100"
    ws_analytics['C6'].number_format = '0.00"%"'

    ws_analytics.column_dimensions['A'].width = 20
    ws_analytics.column_dimensions['B'].width = 15
    ws_analytics.column_dimensions['C'].width = 15

    # Сохраняем в временный файл
    temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
    excel_path = temp_file.name
    temp_file.close()

    wb.save(excel_path)

    return excel_path
